from openpyxl import load_workbook
from re import compile, match, findall
from netmiko import ConnectHandler
from mac_vendor_lookup import MacLookup
from time import strftime


def convert_port_length_in_show_output(any_show_output):
    for j in any_show_output:
        j['port'] = shorten_port(j['port'])
    return any_show_output


def shorten_port(port):
    return ''.join(findall('^.{2}|\d|/', port))


# example of list_of_dict
# list_of_dict = [
# {'key_a': 'value_a', 'key_b': 'value_b'},
# {'key_a': 'value_aa', 'key_b': 'value_bb'},
# {'key_a': 'value_aaa', 'key_b': 'value_bbb'},
# ]
def search_for_someones_attr(list_of_dict, key_a, value_a, key_b):  # returns a single-line string
    for d in list_of_dict:
        if d[key_a] == value_a:
            return d[key_b]


def search_for_someones_multi_attr(list_of_dict, key_a, value_a, key_b):  # returns a multi-line string
    value_b_string_newlines = ''
    for q in list_of_dict:
        if q[key_a] == value_a:
            value_b_string_newlines = f'{value_b_string_newlines}{q[key_b]}\n'
    return value_b_string_newlines[:-1]


def find_endpoint_mac_add_per_vlan(f_show_mac_add, port, the_vlan_given):  # returns a multi-line string
    macs_per_vlan_str_newlines = ''
    for w in f_show_mac_add:
        if w['port'][0] == port and w['vlan'] == the_vlan_given:
            macs_per_vlan_str_newlines = f'{macs_per_vlan_str_newlines}{w["destination_address"]}\n'
    return macs_per_vlan_str_newlines[:-1]


# open up the Inventory.xlsx file that should be pre-filled with information about all the switches
wb_inventory = load_workbook(r'C:\Users\admin\OneDrive - Sydney Airport\Desktop\Python Inputs\Inventory.xlsx')
ws_device_inventory = wb_inventory['Device Inventory']
device_count = ws_device_inventory.max_row - 1

wb_visible = load_workbook(r'C:\Users\admin\OneDrive - Sydney Airport\Desktop\Python Inputs\VISIBLE.xlsx')
ws_main_table = wb_visible['Main Table']

# asw_list is a list of dict to store all the access layer switches
asw_list = []
# dsw_list is a list of dict to store all the distribution layer switches
dsw_list = []

# vlan_db is a dict to store ARP table for each domain --> key will be a str and value will be a list of dict
# vlan_db_dict = {
# 'domain1': [{},{}],
# 'domain2': [{},{}],
# 'domain3': [{},{}],
# }
vlan_db_dict = {}
# arp_table is a type dict to store ARP table for each domain --> key will be a str and value will be a list of dict
arp_table_dict = {}

for i in range(device_count):
    if match('.*-ASW-.*', ws_device_inventory.cell(i + 2, 1).value):
        asw_dictionary = {
            'hostname': ws_device_inventory.cell(i + 2, 1).value,
            'ip_address': ws_device_inventory.cell(i + 2, 2).value,
            'domain': ws_device_inventory.cell(i + 2, 3).value,
            'port_count': ws_device_inventory.cell(i + 2, 4).value,
        }
        asw_list.append(asw_dictionary)
    elif match('.*-DSW-.*', ws_device_inventory.cell(i + 2, 1).value):
        dsw_domain = ws_device_inventory.cell(i + 2, 3).value
        dsw_dictionary = {
            'hostname': ws_device_inventory.cell(i + 2, 1).value,
            'ip_address': ws_device_inventory.cell(i + 2, 2).value,
            'domain': dsw_domain,
        }
        dsw_list.append(dsw_dictionary)
        # building 2 empty dictionaries for ARP Tables and VLAN Databases
        vlan_db_dict.update({dsw_domain: []})
        arp_table_dict.update({dsw_domain: []})

for dsw in dsw_list:
    dsw_connection_parameters = {
        'ip': dsw['ip_address'],
        'device_type': 'cisco_ios',
        'username': 'melghafri',
        'password': 'Airport321!',
    }
    # connect to the DSW and print a connection success message
    dsw_ssh_session = ConnectHandler(**dsw_connection_parameters)
    dsw_hostname = dsw_ssh_session.find_prompt()[:-1]
    print(f'======= CONNECTED TO {dsw_hostname} in {dsw["domain"]} =======')

    # get VALN database from ANY of the 2 DSWs, assuming both DSW switches are VTP server
    vlan_db_dict[dsw['domain']] = dsw_ssh_session.send_command('show vlan brief', use_textfsm=True)

    # get ARP table from BOTH of the 2 DSWs, assuming they are standalone running FHRP
    show_arp_table = dsw_ssh_session.send_command('show ip arp', use_textfsm=True)
    arp_table_dict[dsw['domain']].extend(show_arp_table)

    
k = 2
# taking one switch at a time
for sw in asw_list:
    vlan_db = vlan_db_dict[sw['domain']]
    arp_table = arp_table_dict[sw['domain']]
    sw_connection_parameters = {
        'ip': sw['ip_address'],
        'device_type': 'cisco_ios',
        'username': 'loopbacks',
        'password': 'H3lpm30b!tw@',
    }
    try:
        ssh_session = ConnectHandler(**sw_connection_parameters)
        sw_hostname = ssh_session.find_prompt()[:-1]
        print(f'======= CONNECTED TO {sw_hostname} =======')
        # show interfaces status & building sw_port_list list of dictionaries
        show_int_status = ssh_session.send_command('show interfaces status', use_textfsm=True)
        sw_port_list = []
        # filtering only copper ports by using REGEX to match 'TX' in the port type
        for p in show_int_status:
            if match('.*TX$', p['type']):
                port_dictionary = {
                    'port': p['port'],
                    'vlan': p['vlan'],
                    'status': p['status'],
                }
                sw_port_list.append(port_dictionary)
        # expanding sw_port_list
        pre_show_int = ssh_session.send_command('show interface', use_textfsm=True)
        show_int = convert_port_length_in_show_output(pre_show_int)
        show_int_switchport = ssh_session.send_command("show interfaces switchport", use_textfsm=True)
        show_mac_add = ssh_session.send_command('show mac address-table', use_textfsm=True)
        pre_show_cdp_nei = ssh_session.send_command("show cdp neighbor detail", use_textfsm=True)
        show_cdp_nei = convert_port_length_in_show_output(pre_show_cdp_nei)
        try:
            show_power_inline = ssh_session.send_command('show power inline', use_textfsm=True)
            if type(show_power_inline) == str:
                raise Exception("switch doesn't have PoE")
        except Exception as NoPoE:
            print(NoPoE)
            show_power_inline = []
        for p in sw_port_list:
            p['description'] = search_for_someones_attr(show_int, 'port', p['port'], 'description')
            p['last_input'] = search_for_someones_attr(show_int, 'port', p['port'], 'last_input')
            p['last_output'] = search_for_someones_attr(show_int, 'port', p['port'], 'last_output')
            p['mode'] = search_for_someones_attr(show_int_switchport, 'port', p['port'], 'mode')
            p['trunking_vlans'] = search_for_someones_attr(show_int_switchport, 'port', p['port'], 'trunking_vlans')
            if 'trunk' in p['vlan']:
                # VLANs
                vlan_str = p['trunking_vlans'][0]
                vlan_str_in_newlines = vlan_str.split(',')
                vlan_list_of_int = []
                for h in vlan_str_in_newlines:
                    if '-' in h:
                        unpack = h.split('-')
                        start = int(unpack[0])
                        finish = int(unpack[1]) + 1
                        for n in range(start, finish):
                            vlan_list_of_int.append(str(n))
                    else:
                        vlan_list_of_int.append(h)
                p['vlan'] = '\n'.join(vlan_list_of_int)
                vlan_name_str_in_newlines = ''
                for v in vlan_list_of_int:
                    for vlan in dsw_vlan_db:
                        if vlan['vlan_id'] == v:
                            vlan_name_str_in_newlines = vlan_name_str_in_newlines + vlan['name'] + '\n'
                p['vlan_name'] = vlan_name_str_in_newlines
                p['voice_vlan'] = ''
                p['voice_vlan_name'] = ''

                # Building MAC string variable
                mac_all_vlan = ''
                for v in vlan_list_of_int:
                    mac_per_vlan = find_endpoint_mac_add_per_vlan(show_mac_add, p['port'], str(v))
                    mac_all_vlan = mac_all_vlan + mac_per_vlan + '\n'
                mac_str_in_newlines = mac_all_vlan[:-1]
            else:
                # VLAN
                vlan = p['vlan']
                try:
                    p['vlan'] = int(vlan)
                    p['vlan_name'] = search_for_someones_attr(dsw_vlan_db, 'vlan_id', vlan, 'name')
                except:
                    p['vlan'] = vlan
                    p['vlan_name'] = ''

                voice_vlan = search_for_someones_attr(show_int_switchport, 'port', p['port'], 'voice_vlan')
                try:
                    p['voice_vlan'] = int(voice_vlan)
                    p['voice_vlan_name'] = search_for_someones_attr(dsw_vlan_db, 'vlan_id', voice_vlan, 'name')
                except:
                    p['voice_vlan'] = ''
                    p['voice_vlan_name'] = ''

                # MAC string variable
                mac_str_in_newlines = find_endpoint_mac_add_per_vlan(show_mac_add, p['port'], str(p['vlan']))
            p['mac_learned'] = mac_str_in_newlines
            mac_list = mac_str_in_newlines.splitlines()
            vendor_str_in_newlines = ''
            ip_str_in_newlines = ''
            for m in mac_list:
                try:
                    vendor_str_in_newlines = vendor_str_in_newlines + MacLookup().lookup(m) + '\n'
                except:
                    vendor_str_in_newlines = vendor_str_in_newlines + '!!!!!!!!' + '\n'
                try:
                    ip_str_in_newlines = ip_str_in_newlines + search_for_someones_attr(dsw_arp_table, 'mac', m, 'address') + '\n '
                except:
                    ip_str_in_newlines = ip_str_in_newlines + '!!!!!!!!' + '\n'
            p['mac_lookup'] = vendor_str_in_newlines[:-1]
            p['ip_lookup'] = ip_str_in_newlines[:-1]
            p['cdp_nei_name'] = search_for_someones_multi_attr(show_cdp_nei, 'port', p['port'], 'destination_host')
            p['cdp_nei_platform'] = search_for_someones_multi_attr(show_cdp_nei, 'port', p['port'], 'platform')
            p['cdp_nei_ip'] = search_for_someones_multi_attr(show_cdp_nei, 'port', p['port'], 'management_ip')
            p['power'] = search_for_someones_attr(show_power_inline, 'port', p['port'], 'power')
        # building the Excel sheet rows
        for p in sw_port_list:
            # Col A: switch hostname
            ws_main_table.cell(k, 1).value = sw_hostname
            # Col B: switch domain
            ws_main_table.cell(k, 2).value = sw_domain
            # Col C: port number
            ws_main_table.cell(k, 3).value = p['port']
            # Col D: port description
            ws_main_table.cell(k, 4).value = p['description']
            # Col E: port status
            ws_main_table.cell(k, 5).value = p['status']
            # Col F: switch port mode
            ws_main_table.cell(k, 6).value = p['mode']
            # Col G: VLAN
            ws_main_table.cell(k, 7).value = p['vlan']
            # Col H: VLAN name(s)
            ws_main_table.cell(k, 8).value = p['vlan_name']
            # Col I: Voice VLAN
            ws_main_table.cell(k, 9).value = p['voice_vlan']
            # Col J: Voice VLAN name
            ws_main_table.cell(k, 10).value = p['voice_vlan_name']
            # Col K: MAC learned
            ws_main_table.cell(k, 11).value = p['mac_learned']
            # Col L: Manufacturer(s)
            ws_main_table.cell(k, 12).value = p['mac_lookup']
            # Col M: IP(s) resolved
            ws_main_table.cell(k, 13).value = p['ip_lookup']
            # Col N: CDP neighbor name(s)
            ws_main_table.cell(k, 14).value = p['cdp_nei_name']
            # Col O: CDP neighbor model(s)
            ws_main_table.cell(k, 15).value = p['cdp_nei_platform']
            # Col P: CDP neighbor management IP(s)
            ws_main_table.cell(k, 16).value = p['cdp_nei_ip']
            # Col Q: Last input(s)
            ws_main_table.cell(k, 17).value = p['last_input']
            # Col R: Last output(s)
            ws_main_table.cell(k, 18).value = p['last_output']
            # Col S: Power admin status
            ws_main_table.cell(k, 19).value = p['power']
            k = k + 1
    except Exception as failedConn:
        print(f'!!!!!!! CONNECTION FAILED TO {sw["hostname"]} in {sw["domain"]} !!!!!!!')
        print(failedConn)
        k = k + 1
date_time = strftime('%Y-%m-%d %H-%M')
wb_port2excel.save(rf'Port 2 Excel run at {date_time}.xlsx')
